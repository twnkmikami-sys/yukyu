
import streamlit as st

import hashlib
import pandas as pd
import sqlite3
import zipfile
import os



from datetime import datetime, timedelta

from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill

from openpyxl.worksheet.page import PageMargins


from openpyxl import Workbook
from openpyxl.styles import Font

from openpyxl import Workbook

from openpyxl.styles import (

    Font,
    PatternFill,
    Alignment,
    Border,
    Side

)

from openpyxl.utils import get_column_letter




# =====================================
# DB接続
# =====================================

conn = sqlite3.connect(
    "leave.db",
    check_same_thread=False
)

cursor = conn.cursor()

# =====================================
# テーブル作成
# =====================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS employees (

    employee_id TEXT,

    name TEXT,

    kana TEXT,

    hire_date TEXT,

    branch TEXT

)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS leave_data (

    employee_id TEXT,

    name TEXT,

    granted_days REAL,

    used_days REAL,

    remain_days REAL

)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS leave_history (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    employee_id TEXT,

    name TEXT,

    leave_date TEXT,

    leave_days REAL

)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS attendance (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    employee_id TEXT,

    name TEXT,

    work_date TEXT,

    attendance_type TEXT,

    attendance_days REAL

)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS grant_history (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    employee_id TEXT,

    name TEXT,

    grant_date TEXT,

    grant_days REAL,

    expire_date TEXT,

    used_days REAL,

    expired_flag TEXT

)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS auto_grants (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    employee_id TEXT,

    grant_year REAL,

    grant_date TEXT

)
""")

conn.commit()

# =====================================
# users テーブル
# =====================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS users (

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    username TEXT,

    password TEXT,

    role TEXT,

    employee_id TEXT

)
""")

conn.commit()

# =====================================
# タイトル
# =====================================

# =====================================
# ログイン
# =====================================

st.title("ログイン")

login_user = st.text_input(
    "ユーザー名",
    key="login_user"
)

login_pass = st.text_input(
    "パスワード",
    type="password",
    key="login_pass"
)

# =====================================
# ログイン処理
# =====================================

if st.button("ログイン"):

    hash_pass = hashlib.sha256(
        login_pass.encode()
    ).hexdigest()

    user_df = pd.read_sql(
        f"""
        SELECT * FROM users
        WHERE username='{login_user}'
        AND password='{hash_pass}'
        """,
        conn
    )

    if len(user_df) > 0:

        st.session_state["login"] = True

        st.session_state["role"] = (
            user_df.iloc[0]["role"]
        )

        st.session_state["employee_id"] = (
            user_df.iloc[0]["employee_id"]
        )

        st.success("ログイン成功")

    else:

        st.error(
            "ログイン失敗"
        )

# =====================================
# 未ログイン
# =====================================

if "login" not in st.session_state:

    st.stop()



st.title("有給管理システム")


# =====================================
# ログアウト
# =====================================

col1, col2 = st.columns(2)

with col2:

    if st.button(
        "ログアウト",
        key="logout_button"
    ):

        st.session_state.clear()

        st.success(
            "ログアウトしました"
        )

        st.rerun()

# =====================================
# パスワード変更
# =====================================

st.header("パスワード変更")

current_pass = st.text_input(

    "現在のパスワード",

    type="password",

    key="current_pass"

)

new_pass = st.text_input(

    "新しいパスワード",

    type="password",

    key="new_pass"

)

new_pass2 = st.text_input(

    "新しいパスワード（確認）",

    type="password",

    key="new_pass2"

)

if st.button(

    "パスワード変更",

    key="change_password"

):

    # =====================================
    # 一致確認
    # =====================================

    if new_pass != new_pass2:

        st.error(
            "新しいパスワード不一致"
        )

    else:

        current_hash = hashlib.sha256(

            current_pass.encode()

        ).hexdigest()

        # =====================================
        # 現在ユーザー確認
        # =====================================

        current_user_df = pd.read_sql(
            f"""
            SELECT * FROM users
            WHERE employee_id='{st.session_state["employee_id"]}'
            """,
            conn
        )

        if len(current_user_df) == 0:

            st.error(
                "ユーザー情報なし"
            )

        else:

            db_pass = current_user_df.iloc[0][
                "password"
            ]

            # =====================================
            # 現在PASS確認
            # =====================================

            if current_hash != db_pass:

                st.error(
                    "現在パスワード違います"
                )

            else:

                new_hash = hashlib.sha256(

                    new_pass.encode()

                ).hexdigest()

                cursor.execute("""
                UPDATE users
                SET password=?
                WHERE employee_id=?
                """, (

                    new_hash,

                    st.session_state[
                        "employee_id"
                    ]

                ))

                conn.commit()

                st.success(
                    "パスワード変更完了"
                )
 


# =====================================
# 社員追加
# =====================================

st.header("社員追加")

if "form_key" not in st.session_state:

    st.session_state["form_key"] = 0

with st.form(

    f"employee_form_{
        st.session_state['form_key']
    }"

):


    employee_id_input = st.text_input(
        "社員ID"
    )

    name_input = st.text_input(
        "氏名"
    )

    kana_input = st.text_input(
        "フリガナ"
    )

    hire_date_input = st.date_input(
        "入社日"
    )

    branch_input = st.text_input(
        "所属"
    )

    submit_employee = st.form_submit_button(
        "社員追加"
    )

# =====================================
# 社員追加処理
# =====================================

if submit_employee:

    employee_id_input = (
        employee_id_input.strip()
    )

    name_input = (
        name_input.strip()
    )

    kana_input = (
        kana_input.strip()
    )

    branch_input = (
        branch_input.strip()
    )

    # =====================================
    # 社員ID重複チェック
    # =====================================

    check_employee = pd.read_sql(

        """
        SELECT * FROM employees
        WHERE TRIM(employee_id)=?
        """,

        conn,

        params=(employee_id_input,)

    )

    if len(check_employee) > 0:

        st.error(
            "社員ID重複"
        )

    else:

        # =====================================
        # employees登録
        # =====================================

        cursor.execute("""

        INSERT INTO employees (

            employee_id,
            name,
            kana,
            hire_date,
            branch

        )

        VALUES (?, ?, ?, ?, ?)

        """, (

            employee_id_input,
            name_input,
            kana_input,
            str(hire_date_input),
            branch_input

        ))

        # =====================================
        # leave_data登録
        # =====================================

        cursor.execute("""

        INSERT INTO leave_data (

            employee_id,
            name,
            granted_days,
            used_days,
            remain_days

                       
        )

        VALUES (?, ?, ?, ?, ?)

        """, (

            employee_id_input,
            name_input,
            0,
            0,
            0

        ))

        # =====================================
        # 保存確定
        # =====================================

        conn.commit()

        st.success(
            "社員追加完了"
        )

        st.rerun()

# =====================================
# 社員一覧取得
# =====================================

master_df = pd.read_sql(

    "SELECT * FROM employees",

    conn

)

# =====================================
# 社員一覧表示
# =====================================

st.header("社員一覧")

st.dataframe(master_df)

# =====================================
# 社員選択
# =====================================

employee_list = (

    master_df["name"]

    .dropna()

    .unique()

    .tolist()

)

if len(employee_list) == 0:

    st.warning(
        "社員が存在しません"
    )

    st.stop()

selected_employee = st.selectbox(

    "社員を選択してください",

    employee_list

)

# =====================================
# 選択社員情報
# =====================================

employee_row = master_df[

    master_df["name"]
    == selected_employee

]

employee_id = employee_row.iloc[0][
    "employee_id"
]

# =====================================
# 有給情報取得
# =====================================

leave_df = pd.read_sql(

    "SELECT * FROM leave_data",

    conn

)

employee_leave = leave_df[

    leave_df["employee_id"]
    == employee_id

]

# =====================================
# 社員情報
# =====================================

st.header("社員情報")

st.dataframe(employee_row)

# =====================================
# 有給情報
# =====================================

st.header("有給情報")

st.dataframe(employee_leave)


# =====================================
# 有給取得入力
# =====================================

st.header("有給取得入力")

leave_date = st.date_input(
    "取得日"
)

leave_days = st.number_input(

    "取得日数",

    min_value=0.5,

    max_value=20.0,

    step=0.5,

    value=1.0

)

if st.button("有給登録"):

    # =====================================
    # 履歴登録
    # =====================================

    cursor.execute("""

    INSERT INTO leave_history (

        employee_id,
        name,
        leave_date,
        leave_days

    )

    VALUES (?, ?, ?, ?)

    """, (

        employee_id,
        selected_employee,
        str(leave_date),
        leave_days

    ))

    # =====================================
    # 現在有給取得
    # =====================================

    current_leave = employee_leave.iloc[0]

    new_used = (

        float(current_leave["used_days"])

        + leave_days

    )

    new_remain = (

        float(current_leave["remain_days"])

        - leave_days

    )

    if new_remain < 0:

        new_remain = 0

    # =====================================
    # leave_data更新
    # =====================================

    cursor.execute("""

    UPDATE leave_data

    SET

        used_days=?,

        remain_days=?

    WHERE employee_id=?

    """, (

        new_used,
        new_remain,
        employee_id

    ))

    conn.commit()

    st.success(
        "有給登録完了"
    )

    st.rerun()

# =====================================
# 有給取得履歴
# =====================================

st.header("有給取得履歴")

history_df = pd.read_sql(

    "SELECT * FROM leave_history",

    conn

)

employee_history = history_df[

    history_df["employee_id"]
    == employee_id

]

# 日付順

if len(employee_history) > 0:

    employee_history = (
        employee_history
        .sort_values(
            "leave_date",
            ascending=False
        )
    )

st.dataframe(
    employee_history,
    width="content"
)

# =====================================
# 出勤入力
# =====================================

st.header("出勤入力")

attendance_date = st.date_input(
    "勤務日"
)

attendance_type = st.selectbox(

    "勤務区分",

    [

        "出勤",
        "有給",
        "欠勤",
        "休日"

    ]

)

attendance_days = st.number_input(

    "日数",

    min_value=0.0,

    max_value=1.0,

    step=0.5,

    value=1.0

)

if st.button("出勤登録"):

    cursor.execute("""

    INSERT INTO attendance (

        employee_id,
        name,
        work_date,
        attendance_type,
        attendance_days

    )

    VALUES (?, ?, ?, ?, ?)

    """, (

        employee_id,
        selected_employee,
        str(attendance_date),
        attendance_type,
        attendance_days

    ))

    conn.commit()

    st.success(
        "出勤登録完了"
    )

    st.rerun()

# =====================================
# 出勤履歴
# =====================================

st.header("出勤履歴")

attendance_df = pd.read_sql(

    "SELECT * FROM attendance",

    conn

)

employee_attendance = attendance_df[

    attendance_df["employee_id"]
    == employee_id

]

if len(employee_attendance) > 0:

    employee_attendance = (
        employee_attendance
        .sort_values(
            "work_date",
            ascending=False
        )
    )

st.dataframe(

    employee_attendance,

    width="content"

)

# =====================================
# 月別集計
# =====================================

st.header("月別集計")

if len(employee_attendance) > 0:

    # 日付変換

    employee_attendance["work_date"] = (
        pd.to_datetime(
            employee_attendance["work_date"]
        )
    )

    # 年月列

    employee_attendance["year_month"] = (

        employee_attendance["work_date"]

        .dt.strftime("%Y-%m")

    )

    # 集計

    summary_df = employee_attendance.groupby(

        [

            "year_month",

            "attendance_type"

        ]

    )["attendance_days"].sum().reset_index()

    # ピボット

    pivot_df = summary_df.pivot_table(

        index="year_month",

        columns="attendance_type",

        values="attendance_days",

        aggfunc="sum",

        fill_value=0

    ).reset_index()

    # 列不足対策

    for col in [

        "出勤",

        "有給",

        "欠勤",

        "休日"

    ]:

        if col not in pivot_df.columns:

            pivot_df[col] = 0

    # 出勤扱い

    pivot_df["出勤扱い"] = (

        pivot_df["出勤"]

        + pivot_df["有給"]

    )

    # 総日数

    pivot_df["総日数"] = (

        pivot_df["出勤"]

        + pivot_df["有給"]

        + pivot_df["欠勤"]

    )

    # 出勤率

    pivot_df["出勤率"] = 0

    pivot_df.loc[

        pivot_df["総日数"] > 0,

        "出勤率"

    ] = (

        pivot_df["出勤扱い"]

        / pivot_df["総日数"]

    ) * 100

    pivot_df["出勤率"] = (

        pivot_df["出勤率"]

        .round(1)

    )

    st.dataframe(

        pivot_df,

        width="content"
    )

# =====================================
# 8割判定
# =====================================

st.header("8割判定")

if len(employee_attendance) > 0:

    # 全期間出勤率

    work_days = employee_attendance[

        employee_attendance["attendance_type"]

        .isin([

            "出勤",

            "有給"

        ])

    ]["attendance_days"].sum()

    total_days = employee_attendance[

        employee_attendance["attendance_type"]

        != "休日"

    ]["attendance_days"].sum()

    attendance_rate = 0

    if total_days > 0:

        attendance_rate = (

            work_days
            / total_days

        ) * 100

    attendance_rate = round(
        attendance_rate,
        1
    )

    # 判定

    if attendance_rate >= 80:

        st.success(

            f"""
            出勤率:
            {attendance_rate}%
            （8割達成）
            """

        )

    else:

        st.error(

            f"""
            出勤率:
            {attendance_rate}%
            （8割未達）
            """

        )

else:

    st.info(
        "出勤データなし"
    )
# =====================================
# 年5日取得義務
# =====================================

st.header("年5日取得義務チェック")

current_year = datetime.today().year

# =====================================
# 有給履歴なし
# =====================================

if len(history_df) == 0:

    st.info(
        "有給取得履歴なし"
    )

else:

    # 日付変換

    history_df["leave_date"] = (
        pd.to_datetime(
            history_df["leave_date"]
        )
    )

    # 当年取得

    year_history = history_df[

        history_df["leave_date"]

        .dt.year

        == current_year

    ]

    # 社員別集計

    year_summary = year_history.groupby(

        "employee_id"

    )["leave_days"].sum().reset_index()

    warning_list = []

    # =====================================
    # 全社員確認
    # =====================================

    for index, row in leave_df.iterrows():

        emp_id = row["employee_id"]

        emp_name = row["name"]

        granted = float(
            row["granted_days"]
        )

        # 10日以上付与対象

        if granted >= 10:

            emp_history = year_summary[

                year_summary["employee_id"]
                == emp_id

            ]

            used_days = 0

            if len(emp_history) > 0:

                used_days = float(

                    emp_history.iloc[0][
                        "leave_days"
                    ]

                )

            remain_need = max(

                0,

                5 - used_days

            )

            status = (

                "達成"

                if used_days >= 5

                else "未達"

            )

            warning_list.append({

                "社員ID": emp_id,

                "氏名": emp_name,

                "付与日数": granted,

                "取得日数": used_days,

                "不足日数": remain_need,

                "状態": status

            })

    # DataFrame化

    warning_df = pd.DataFrame(
        warning_list
    )

    # =====================================
    # 対象なし
    # =====================================

    if len(warning_df) == 0:

        st.info(
            "対象社員なし"
        )

    else:

        st.dataframe(

            warning_df,

            width="content"

        )

        # =====================================
        # 未達一覧
        # =====================================

        not_clear_df = warning_df[

            warning_df["状態"]
            == "未達"

        ]

        if len(not_clear_df) > 0:

            st.warning(

                f"""
                年5日未達者:
                {len(not_clear_df)}名
                """

            )

        else:

            st.success(
                "全員達成"
            )
            
# =====================================
# 消滅予定一覧
# =====================================

st.header("消滅予定一覧")

grant_df = pd.read_sql(

    "SELECT * FROM grant_history",

    conn

)

# =====================================
# 社員別
# =====================================

employee_grants = grant_df[

    grant_df["employee_id"]
    == employee_id

]

if len(employee_grants) == 0:

    st.info(
        "付与履歴なし"
    )

else:

    today_date = datetime.today().date()

    expire_list = []

    for index, row in employee_grants.iterrows():

        expire_date = datetime.strptime(

            row["expire_date"],

            "%Y-%m-%d"

        ).date()

        remain_days = (

            float(row["grant_days"])

            - float(row["used_days"])

        )

        if remain_days < 0:

            remain_days = 0

        # 残り日数

        left_days = (

            expire_date
            - today_date

        ).days

        # 状態

        if left_days < 0:

            status = "消滅済"

        elif left_days <= 30:

            status = "1か月以内"

        else:

            status = "正常"

        expire_list.append({

            "付与日": row["grant_date"],

            "消滅日": row["expire_date"],

            "残日数": remain_days,

            "残り日数": left_days,

            "状態": status

        })

    expire_df = pd.DataFrame(
        expire_list
    )

    st.dataframe(

        expire_df,

        width="content"

    )

    # =====================================
    # 警告表示
    # =====================================

    warning_df = expire_df[

        expire_df["状態"]
        == "1か月以内"

    ]

    if len(warning_df) > 0:

        st.warning(

            f"""
            消滅1か月以内:
            {len(warning_df)}件
            """

        )

    expired_df = expire_df[

        expire_df["状態"]
        == "消滅済"

    ]

    if len(expired_df) > 0:

        st.error(

            f"""
            消滅済:
            {len(expired_df)}件
            """

        )
        
# =====================================
# メインメニュー
# =====================================
st.sidebar.title("メニュー")
menu = st.sidebar.selectbox(
    "機能を選択",
    ["Excel個人台帳出力", "全社員Excel一括出力"]
)

        
# =====================================
# Excel個人台帳出力（プレビュー付き）
# =====================================
if menu == "Excel個人台帳出力":
    st.header("Excel個人台帳出力")

    # 状態を記憶：データ保持用
    if "excel_data" not in st.session_state:
        st.session_state["excel_data"] = None
    if "excel_filename" not in st.session_state:
        st.session_state["excel_filename"] = None
    if "preview_personal" not in st.session_state:
        st.session_state["preview_personal"] = None

    # 社員選択
    emp_list_df = pd.read_sql("SELECT employee_id, name FROM employees", conn)
    emp_list = emp_list_df["name"].tolist()
    selected_employee = st.selectbox("社員を選択", emp_list)
    employee_id = emp_list_df[emp_list_df["name"] == selected_employee]["employee_id"].iloc[0]

    if st.button("Excel作成", key="excel_output"):
        # 各種データ取得
        employee_row = pd.read_sql("SELECT * FROM employees WHERE employee_id=?", conn, params=(employee_id,))
        employee_leave = pd.read_sql("SELECT * FROM leave_data WHERE employee_id=?", conn, params=(employee_id,))
        employee_history = pd.read_sql("SELECT * FROM leave_history WHERE employee_id=?", conn, params=(employee_id,))
        employee_attendance = pd.read_sql("SELECT * FROM attendance WHERE employee_id=?", conn, params=(employee_id,))
        grant_export = pd.read_sql("SELECT * FROM grant_history WHERE employee_id=?", conn, params=(employee_id,))

        # Workbook 作成
        wb = Workbook()
        ws = wb.active
        ws.title = "有給台帳"

        # 色設定
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        # 印刷設定
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

        # タイトル
        ws.merge_cells("A1:E1")
        ws["A1"] = "有給休暇管理台帳"
        ws["A1"].font = Font(bold=True, size=18)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 社員情報
        ws["A3"] = "社員ID"
        ws["B3"] = employee_id
        ws["A4"] = "氏名"
        ws["B4"] = selected_employee
        ws["A5"] = "所属"
        ws["B5"] = employee_row.iloc[0]["branch"]
        ws["A6"] = "入社日"
        ws["B6"] = employee_row.iloc[0]["hire_date"]

        # 有給情報
        leave_row = employee_leave.iloc[0]
        ws["D3"] = "付与日数"
        ws["E3"] = float(leave_row["granted_days"])
        ws["D4"] = "使用日数"
        ws["E4"] = float(leave_row["used_days"])
        ws["D5"] = "残日数"
        ws["E5"] = float(leave_row["remain_days"])

        # 出勤率計算
        attendance_rate = 0
        if len(employee_attendance) > 0:
            work_days = employee_attendance[
                employee_attendance["attendance_type"].isin(["出勤", "有給"])
            ]["attendance_days"].sum()
            total_days = employee_attendance[
                employee_attendance["attendance_type"] != "休日"
            ]["attendance_days"].sum()
            if total_days > 0:
                attendance_rate = round(work_days / total_days * 100, 1)

        ws["D9"] = "出勤率"
        ws["E9"] = f"{attendance_rate}%"
        if attendance_rate >= 80:
            ws["E9"].fill = green_fill
        else:
            ws["E9"].fill = red_fill

        # 有給履歴
        ws["A9"] = "取得日"
        ws["B9"] = "取得日数"
        row_no = 10
        history_data = []
        for _, row in employee_history.iterrows():
            ws.cell(row=row_no, column=1).value = row["leave_date"]
            ws.cell(row=row_no, column=2).value = row["leave_days"]
            history_data.append({"取得日": row["leave_date"], "取得日数": row["leave_days"]})
            row_no += 1

        # 消滅警告
        warning_list = []
        warning_row = max(row_no + 2, 15)
        ws[f"D{warning_row}"] = "消滅予定"
        today_date = datetime.today().date()
        warning_row += 1
        for _, row in grant_export.iterrows():
            if row["expired_flag"] == "消滅済":
                continue
            expire_date = datetime.strptime(row["expire_date"], "%Y-%m-%d").date()
            remain_days = float(row["grant_days"]) - float(row["used_days"])
            if remain_days <= 0:
                continue
            days_left = (expire_date - today_date).days
            if days_left <= 30:
                ws[f"D{warning_row}"] = "期限接近"
                ws[f"E{warning_row}"] = str(expire_date)
                ws[f"F{warning_row}"] = remain_days
                ws[f"D{warning_row}"].fill = yellow_fill
                ws[f"E{warning_row}"].fill = yellow_fill
                ws[f"F{warning_row}"].fill = yellow_fill
                warning_list.append({"状態": "期限接近", "期限日": expire_date, "残日数": remain_days})
                warning_row += 1

        # 列幅
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15

        # 枠線
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # --- プレビュー用データを保存 ---
        st.session_state["preview_personal"] = {
            "社員情報": [
                {"項目": "社員ID", "内容": employee_id},
                {"項目": "氏名", "内容": selected_employee},
                {"項目": "所属", "内容": employee_row.iloc[0]["branch"]},
                {"項目": "入社日", "内容": employee_row.iloc[0]["hire_date"]}
            ],
            "有給情報": [
                {"項目": "付与日数", "内容": float(leave_row["granted_days"])},
                {"項目": "使用日数", "内容": float(leave_row["used_days"])},
                {"項目": "残日数", "内容": float(leave_row["remain_days"])},
                {"項目": "出勤率", "内容": f"{attendance_rate}%"}
            ],
            "有給取得履歴": pd.DataFrame(history_data),
            "消滅警告リスト": pd.DataFrame(warning_list)
        }

        # Excelデータをメモリ保存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        st.session_state["excel_data"] = excel_buffer
        st.session_state["excel_filename"] = f"{selected_employee}_有給台帳.xlsx"
        st.success("✅ Excel作成完了　→　下で内容を確認し、ダウンロードできます")

    # --- プレビュー表示エリア ---
    if st.session_state["preview_personal"] is not None:
        st.subheader("👀 出力内容プレビュー")
        col1, col2 = st.columns(2)
        with col1:
            st.write("**社員情報**")
            st.dataframe(pd.DataFrame(st.session_state["preview_personal"]["社員情報"]), hide_index=True)
        with col2:
            st.write("**有給・勤怠情報**")
            st.dataframe(pd.DataFrame(st.session_state["preview_personal"]["有給情報"]), hide_index=True)

        st.write("**有給取得履歴**")
        st.dataframe(st.session_state["preview_personal"]["有給取得履歴"], hide_index=True, use_container_width=True)

        if len(st.session_state["preview_personal"]["消滅警告リスト"]) > 0:
            st.warning("**⚠️ 消滅期限が近い有給**")
            st.dataframe(st.session_state["preview_personal"]["消滅警告リスト"], hide_index=True)

    # --- ダウンロードボタン ---
    if st.session_state["excel_data"] is not None:
        st.download_button(
            label="📥 個人台帳ダウンロード",
            data=st.session_state["excel_data"],
            file_name=st.session_state["excel_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# =====================================
# 全社員Excel一括出力（プレビュー＆ZIPダウンロード付き）
# =====================================
if menu == "全社員Excel一括出力":
    st.header("全社員Excel一括出力")

    # 状態記憶
    if "all_zip_data" not in st.session_state:
        st.session_state["all_zip_data"] = None
    if "all_zip_filename" not in st.session_state:
        st.session_state["all_zip_filename"] = None
    if "preview_all" not in st.session_state:
        st.session_state["preview_all"] = None

    if st.button("全社員Excel出力", key="all_individual_excel"):
        zip_buffer = io.BytesIO()
        preview_summary = [] # 一覧プレビュー用

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            all_employee_df = pd.read_sql("SELECT * FROM employees", conn)

            for idx, emp_row in all_employee_df.iterrows():
                emp_id = emp_row["employee_id"]
                emp_name = emp_row["name"]
                emp_branch = emp_row["branch"]
                emp_hire = emp_row["hire_date"]

                # データ取得
                leave_data_df = pd.read_sql("SELECT * FROM leave_data WHERE employee_id=?", conn, params=(emp_id,))
                if len(leave_data_df) == 0:
                    continue
                history_df_export = pd.read_sql("SELECT * FROM leave_history WHERE employee_id=?", conn, params=(emp_id,))
                attendance_df = pd.read_sql("SELECT * FROM attendance WHERE employee_id=?", conn, params=(emp_id,))
                grant_df = pd.read_sql("SELECT * FROM grant_history WHERE employee_id=?", conn, params=(emp_id,))

                # 出勤率計算
                attendance_rate = 0
                if len(attendance_df) > 0:
                    work_days = attendance_df[attendance_df["attendance_type"].isin(["出勤", "有給"])]["attendance_days"].sum()
                    total_days = attendance_df[attendance_df["attendance_type"] != "休日"]["attendance_days"].sum()
                    if total_days > 0:
                        attendance_rate = round(work_days / total_days * 100, 1)

                # 有給情報
                leave_row = leave_data_df.iloc[0]
                granted = float(leave_row["granted_days"])
                used = float(leave_row["used_days"])
                remain = float(leave_row["remain_days"])

                # プレビュー用サマリー追加
                preview_summary.append({
                    "社員ID": emp_id,
                    "氏名": emp_name,
                    "所属": emp_branch,
                    "付与日数": granted,
                    "使用日数": used,
                    "残日数": remain,
                    "出勤率": f"{attendance_rate}%",
                    "状態": "要注意" if attendance_rate < 80 else "通常"
                })

                # --- Excelファイル作成（個人台


    # =====================================
    # 空白除去
    # =====================================

employee_id_input = (
        employee_id_input.strip()
    )

name_input = (
        name_input.strip()
    )

kana_input = (
        kana_input.strip()
    )

branch_input = (
        branch_input.strip()
    )
cursor.execute("""

        INSERT INTO leave_data (

            employee_id,
            name,
            granted_days,
            used_days,
            remain_days

        )

        VALUES (?, ?, ?, ?, ?)

        """, (

            employee_id_input,
            name_input,
            0,
            0,
            0

        ))

conn.commit()

st.success(
            "社員追加完了"
        )

st.rerun()

        # =====================================
        # フォームリセット
        # =====================================

st.session_state["form_key"] += 1

st.rerun()


# =====================================
# 社員取得
# =====================================

master_df = pd.read_sql(
    "SELECT * FROM employees",
    conn
)

if len(master_df) == 0:

    st.warning(
        "社員が登録されていません"
    )

    st.stop()

 # =====================================
# 社員削除
# =====================================

if st.session_state["role"] == "admin":

    st.header("社員削除")

    delete_employee = st.selectbox(

        "削除社員",

        master_df["name"].tolist(),

        key="delete_employee"

    )

    if st.button(
        "社員削除",
        key="delete_employee_button"
    ):

        delete_row = master_df[
            master_df["name"]
            == delete_employee
        ]

        delete_id = delete_row.iloc[0][
            "employee_id"
        ]

        # employees
        cursor.execute("""
        DELETE FROM employees
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        # leave_data
        cursor.execute("""
        DELETE FROM leave_data
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        # leave_history
        cursor.execute("""
        DELETE FROM leave_history
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        # attendance
        cursor.execute("""
        DELETE FROM attendance
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        # grant_history
        cursor.execute("""
        DELETE FROM grant_history
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        # auto_grants
        cursor.execute("""
        DELETE FROM auto_grants
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        # users
        cursor.execute("""
        DELETE FROM users
        WHERE employee_id=?
        """, (

            delete_id,

        ))

        conn.commit()

        st.success(
            "社員削除完了"
        )

        st.rerun()      

# =====================================
# 管理者ダッシュボード
# =====================================

if st.session_state["role"] == "admin":

    st.header("管理者ダッシュボード")

    # =====================================
    # ダッシュボード用DB読込
    # =====================================

    leave_df_dashboard = pd.read_sql(
        "SELECT * FROM leave_data",
        conn
    )

    attendance_dashboard = pd.read_sql(
        "SELECT * FROM attendance",
        conn
    )



    # =====================================
    # 社員数
    # =====================================

    employee_count = len(master_df)

    # =====================================
    # 有給残数
    # =====================================
    total_leave = leave_df_dashboard[
        "remain_days"
     ].sum()

    # =====================================
    # 年5日未達者
    # =====================================

    not_clear_count = 0

    if "warning_df" in locals():

        if len(warning_df) > 0:

            not_clear_count = len(

                warning_df[
                    warning_df["状態"]
                    == "未達"
                ]

            )

    # =====================================
    # 今日出勤
    # =====================================

    attendance_df = pd.read_sql(
        "SELECT * FROM attendance",
        conn
    )

    today_str = datetime.today().strftime(
        "%Y-%m-%d"
    )

    today_attendance = attendance_df[

        (
            attendance_df["work_date"]
            == today_str
        )

        &

        (
            attendance_df[
                "attendance_type"
            ]
            == "出勤"
        )

    ]

    today_count = len(
        today_attendance
    )

    # =====================================
    # 消滅予定
    # =====================================

    grant_df = pd.read_sql(
        "SELECT * FROM grant_history",
        conn
    )

    expire_count = 0

    today_date = datetime.today().date()

    for index, row in grant_df.iterrows():

        if row["expired_flag"] == "消滅済":

            continue

        expire_date = datetime.strptime(

            row["expire_date"],

            "%Y-%m-%d"

        ).date()

        remain_days = (

            float(row["grant_days"])

            - float(row["used_days"])

        )

        # 30日以内
        if (

            expire_date
            <= today_date + timedelta(days=30)

            and remain_days > 0

        ):

            expire_count += 1

    # =====================================
    # KPI表示
    # =====================================

    col1, col2, col3 = st.columns(3)

    with col1:

        st.metric(
            "社員数",
            employee_count
        )

        st.metric(
            "今日出勤",
            today_count
        )

    with col2:

        st.metric(
            "有給残数合計",
            total_leave
        )

        st.metric(
            "消滅予定",
            expire_count
        )

    with col3:

        st.metric(
            "年5日未達",
            not_clear_count
        )


# =====================================
# 権限制御
# =====================================

if st.session_state["role"] == "admin":

    view_df = master_df.copy()

else:

    login_emp_id = st.session_state[
        "employee_id"
    ]

    view_df = master_df[
        master_df["employee_id"]
        == login_emp_id
    ]

# =====================================
# ユーザー登録
# =====================================

if st.session_state["role"]== "admin":

    st.header("ユーザー登録")

    new_username = st.text_input(
        "ログインID",
        key="new_username"
    )

    new_password = st.text_input(
        "パスワード",
        type="password",
        key="new_password"
    )

    new_role = st.selectbox(

        "権限",

        [

            "admin",
            "employee"

        ]
,

        key="new_role"

    )

    employee_options = (
        master_df
["employee_id"]
        .tolist()
    )

    selected_emp_id = st.selectbox(

        "社員ID",

        employee_options,

        key="selected_emp_id"

    )

    if st.button(
        "ユーザー作成"
    ):

        # 重複チェック
        check_user = pd.read_sql(
            f"""
            SELECT * FROM users
            WHERE username='{new_username}'
            """,
            conn
        )

        if len(check_user) > 0:

            st.error(
                "ユーザーID重複"
            )

        else:

            hash_pass = hashlib.sha256(
                new_password.encode()
            ).hexdigest()

            cursor.execute("""
            INSERT INTO users (

                username,
                password,
                role,
                employee_id

            )
            VALUES (?, ?, ?, ?)
            """, (

                new_username,
                hash_pass,
                new_role,
                selected_emp_id

            ))

            conn.commit()

            st.success(
                "ユーザー作成完了"
            )

employee_list = (
    view_df["name"]
    .dropna()
    .unique()
    .tolist()
)

# =====================================
# 全社員一覧
# =====================================

if st.session_state["role"] == "admin":

    st.header("全社員一覧")

    # employees
    all_emp_df = pd.read_sql(
        "SELECT * FROM employees",
        conn
    )

    # leave_data
    all_leave_df = pd.read_sql(
        "SELECT * FROM leave_data",
        conn
    )

    # merge
    all_summary = pd.merge(

        all_emp_df,

        all_leave_df,

        on=[

            "employee_id",
            "name"

        ],

        how="left"

    )

    # 表示列
    show_df = all_summary[[

        "employee_id",

        "name",

        "branch",

        "granted_days",

        "used_days",

        "remain_days"

    ]]

    # 列名変更
    show_df.columns = [

        "社員ID",

        "氏名",

        "所属",

        "付与",

        "使用",

        "残数"

    ]

# =====================================
# 年5日未達一覧
# =====================================

if st.session_state["role"] == "admin":

    st.header("年5日未達一覧")

    current_year = datetime.today().year

    # leave_history
    all_history_df = pd.read_sql(
        "SELECT * FROM leave_history",
        conn
    )

    # leave_data
    all_leave_df = pd.read_sql(
        "SELECT * FROM leave_data",
        conn
    )

    if len(all_history_df) > 0:

        all_history_df["leave_date"] = pd.to_datetime(

            all_history_df["leave_date"]

        )

        current_history = all_history_df[

            all_history_df["leave_date"]
            .dt.year
            == current_year

        ]

        history_summary = current_history.groupby(

            [
                "employee_id",
                "name"
            ]

        )["leave_days"].sum().reset_index()

    else:

        history_summary = pd.DataFrame()

    warning_list = []

    for index, row in all_leave_df.iterrows():

        granted_days = float(
            row["granted_days"]
        )

        # 年10日以上付与対象
        if granted_days >= 10:

            emp_id = row["employee_id"]

            emp_name = row["name"]

            used_days = 0

            if len(history_summary) > 0:

                emp_history = history_summary[

                    history_summary[
                        "employee_id"
                    ] == emp_id

                ]

                if len(emp_history) > 0:

                    used_days = float(

                        emp_history.iloc[0][
                            "leave_days"
                        ]

                    )

            remain_need = max(
                0,
                5 - used_days
            )

            if used_days < 5:

                warning_list.append({

                    "社員ID": emp_id,

                    "氏名": emp_name,

                    "取得日数": used_days,

                    "不足日数": remain_need

                })

    warning_df = pd.DataFrame(
        warning_list
    )

    if len(warning_df) == 0:

        st.success(
            "未達社員なし"
        )

    else:

        def warning_color(val):

            return (
                "background-color: #FF9999"
            )

        styled_warning = warning_df.style.map(

            warning_color,

            subset=["不足日数"]

        )

        st.dataframe(

            styled_warning,

            width="content"

        )

# =====================================
# 消滅予定一覧
# =====================================

if st.session_state["role"] == "admin":

    st.header("消滅予定一覧")

    grant_df_all = pd.read_sql(
        "SELECT * FROM grant_history",
        conn
    )

    today_date = datetime.today().date()

    expire_list = []

    for index, row in grant_df_all.iterrows():

        # 消滅済除外
        if row["expired_flag"] == "消滅済":

            continue

        expire_date = datetime.strptime(

            row["expire_date"],

            "%Y-%m-%d"

        ).date()

        remain_days = (

            float(row["grant_days"])
            - float(row["used_days"])

        )

        # 残無し除外
        if remain_days <= 0:

            continue

        days_left = (
            expire_date - today_date
        ).days

        # 90日以内
        if days_left <= 90:

            expire_list.append({

                "氏名": row["name"],

                "消滅日": str(expire_date),

                "残日数": remain_days,

                "残り日数": days_left

            })

    expire_df = pd.DataFrame(
        expire_list
    )

    if len(expire_df) == 0:

        st.success(
            "消滅予定なし"
        )

    else:

        # ソート
        expire_df = expire_df.sort_values(

            "残り日数"

        )

        # 色
        def expire_color(val):

            try:

                if int(val) <= 30:

                    return (
                        "background-color: #FF9999"
                    )

                elif int(val) <= 60:

                    return (
                        "background-color: #FFFF99"
                    )

                else:

                    return ""

            except:

                return ""

        styled_expire = expire_df.style.map(

            expire_color,

            subset=["残り日数"]

        )

        st.dataframe(

            styled_expire,

            width="content"

        )     

    # =====================================
    # 所属絞込
    # =====================================

    branch_list = (
        ["全体"]
        + sorted(
            show_df["所属"]
            .dropna()
            .unique()
            .tolist()
        )
    )

    selected_branch = st.selectbox(

        "所属選択",

        branch_list,

        key="branch_filter"

    )

    if selected_branch != "全体":

        show_df = show_df[

            show_df["所属"]
            == selected_branch

        ]

    # =====================================
    # 検索
    # =====================================

    search_text = st.text_input(

        "社員検索",

        key="employee_search"

    )

    if search_text != "":

        show_df = show_df[

            show_df.apply(

                lambda row:

                    row.astype(str)
                    .str.contains(

                        search_text,

                        case=False

                    ).any(),

                axis=1

            )

        ]


    # =====================================
    # 残数色変更
    # =====================================

    def remain_color(val):

        try:

            if float(val) <= 0:

                return (
                    "background-color: #FF9999"
                )

            elif float(val) <= 3:

                return (
                    "background-color: #FFFF99"
                )

            else:

                return ""

        except:

            return ""

    styled_df = show_df.style.map(

        remain_color,

        subset=["残数"]

    )

    st.dataframe(

        styled_df,

        width="content"
    )




# =====================================
# 社員選択
# =====================================

selected_employee = st.selectbox(
    "社員を選択してください",
    employee_list
)

employee_row = master_df[
    master_df["name"] == selected_employee
]

employee_id = employee_row.iloc[0]["employee_id"]

# =====================================
# 社員情報
# =====================================

st.header("社員情報")

st.dataframe(employee_row)

# =====================================
# 有給情報
# =====================================

st.header("有給情報")

leave_df = pd.read_sql(
    "SELECT * FROM leave_data",
    conn
)

employee_leave = leave_df[
    leave_df["employee_id"] == employee_id
]

st.dataframe(employee_leave)

# =====================================
# 有給取得入力
# =====================================

st.header("有給取得入力")

leave_date = st.date_input(
    "取得日"
)

leave_days = st.number_input(
    "取得日数",
    min_value=0.5,
    max_value=20.0,
    step=0.5
)

if st.button("有給登録"):

    # 履歴登録
    cursor.execute("""
    INSERT INTO leave_history (
        employee_id,
        name,
        leave_date,
        leave_days
    )
    VALUES (?, ?, ?, ?)
    """, (

        employee_id,
        selected_employee,
        str(leave_date),
        leave_days

    ))

    # =====================================
    # 古い有給優先消化
    # =====================================

    remain_use = leave_days

    grant_df = pd.read_sql(
        f"""
        SELECT * FROM grant_history
        WHERE employee_id='{employee_id}'
        AND expired_flag='未消滅'
        ORDER BY grant_date ASC
        """,
        conn
    )

    for index, row in grant_df.iterrows():

        if remain_use <= 0:

            break

        grant_days = float(row["grant_days"])

        used_days = float(row["used_days"])

        remain_days = (
            grant_days - used_days
        )

        if remain_days <= 0:

            continue

        use_now = min(
            remain_days,
            remain_use
        )

        new_used = (
            used_days + use_now
        )

        cursor.execute("""
        UPDATE grant_history
        SET used_days=?
        WHERE id=?
        """, (

            new_used,
            int(row["id"])

        ))

        remain_use -= use_now

    # leave_data更新
    current_leave = employee_leave.iloc[0]

    new_used_total = (
        float(current_leave["used_days"])
        + leave_days
    )

    new_remain_total = (
        float(current_leave["remain_days"])
        - leave_days
    )

    if new_remain_total < 0:

        new_remain_total = 0

    cursor.execute("""
    UPDATE leave_data
    SET used_days=?,
        remain_days=?
    WHERE employee_id=?
    """, (

        new_used_total,
        new_remain_total,
        employee_id

    ))

    conn.commit()

    st.success("有給登録完了")

# =====================================
# 有給取得履歴
# =====================================

st.header("有給取得履歴")

history_df = pd.read_sql(
    "SELECT * FROM leave_history",
    conn
)

employee_history = history_df[
    history_df["employee_id"] == employee_id
]

st.dataframe(employee_history)

# =====================================
# 月別出勤入力
# =====================================

st.header("月別出勤入力")

attendance_date = st.date_input(
    "勤務日"
)

attendance_type = st.selectbox(

    "勤務区分",

    [

        "出勤",

        "有給",

        "欠勤",

        "休日",

        "特休",

        "労災",

        "産休",

        "育休",

        "休業"

    ]

)

attendance_days = st.number_input(
    "日数",
    min_value=0.0,
    max_value=1.0,
    step=0.5,
    value=1.0
)

if st.button("出勤登録"):

    cursor.execute("""
    INSERT INTO attendance (

        employee_id,
        name,

        work_date,

        attendance_type,

        attendance_days

    )
    VALUES (?, ?, ?, ?, ?)
    """, (

        employee_id,
        selected_employee,

        str(attendance_date),

        attendance_type,

        attendance_days

    ))

    conn.commit()

    st.success("出勤登録完了")

# =====================================
# 月別出勤一覧
# =====================================

st.header("月別出勤一覧")

attendance_df = pd.read_sql(
    "SELECT * FROM attendance",
    conn
)

employee_attendance = attendance_df[
    attendance_df["employee_id"] == employee_id
]

st.dataframe(employee_attendance)

# =====================================
# 月別集計
# =====================================

st.header("月別集計")

if len(employee_attendance) > 0:

    employee_attendance["work_date"] = pd.to_datetime(
        employee_attendance["work_date"]
    )

    employee_attendance["year_month"] = (
        employee_attendance["work_date"]
        .dt.strftime("%Y-%m")
    )

    summary_df = employee_attendance.groupby(

        [
            "year_month",
            "attendance_type"
        ]

    )["attendance_days"].sum().reset_index()

    pivot_df = summary_df.pivot_table(

        index="year_month",

        columns="attendance_type",

        values="attendance_days",

        aggfunc="sum",

        fill_value=0

    ).reset_index()

    if "出勤" not in pivot_df.columns:

        pivot_df["出勤"] = 0

    if "有給" not in pivot_df.columns:

        pivot_df["有給"] = 0

    if "欠勤" not in pivot_df.columns:

        pivot_df["欠勤"] = 0

    pivot_df["出勤扱い"] = (
        pivot_df["出勤"]
        + pivot_df["有給"]
    )

    pivot_df["総日数"] = (
        pivot_df["出勤"]
        + pivot_df["有給"]
        + pivot_df["欠勤"]
    )

    pivot_df["出勤率"] = 0

    pivot_df.loc[
        pivot_df["総日数"] > 0,
        "出勤率"
    ] = (

        pivot_df["出勤扱い"]
        / pivot_df["総日数"]

    ) * 100

    pivot_df["出勤率"] = pivot_df[
        "出勤率"
    ].round(1)

    st.dataframe(pivot_df)

# =====================================
# 自動有給付与
# =====================================

st.header("自動有給付与")

today = datetime.today()

hire_date = datetime.strptime(
    employee_row.iloc[0]["hire_date"],
    "%Y-%m-%d"
)

grant_rules = [

    (0.5, 10),
    (1.5, 11),
    (2.5, 12),
    (3.5, 14),
    (4.5, 16),
    (5.5, 18),
    (6.5, 20)

]

for years, grant_days in grant_rules:

    target_date = hire_date + timedelta(
        days=int(years * 365)
    )

    if today >= target_date:

        auto_df = pd.read_sql(
            f"""
            SELECT * FROM auto_grants
            WHERE employee_id='{employee_id}'
            AND grant_year='{years}'
            """,
            conn
        )

        if len(auto_df) == 0:

            attendance_rate = 100

            if len(employee_attendance) > 0:

                work_days = employee_attendance[
                    employee_attendance["attendance_type"]
                    .isin(["出勤", "有給"])
                ]["attendance_days"].sum()

                all_days = employee_attendance[
                    employee_attendance["attendance_type"]
                    != "休日"
                ]["attendance_days"].sum()

                if all_days > 0:

                    attendance_rate = (
                        work_days / all_days
                    ) * 100

            if attendance_rate >= 80:

                current_leave = employee_leave.iloc[0]

                granted_total = (
                    float(current_leave["granted_days"])
                    + grant_days
                )

                remain_total = (
                    float(current_leave["remain_days"])
                    + grant_days
                )

                cursor.execute("""
                UPDATE leave_data
                SET granted_days=?,
                    remain_days=?
                WHERE employee_id=?
                """, (

                    granted_total,
                    remain_total,
                    employee_id

                ))

                expire_date = (
                    target_date + timedelta(days=730)
                )

                cursor.execute("""
                INSERT INTO grant_history (
                    employee_id,
                    name,
                    grant_date,
                    grant_days,
                    expire_date,
                    used_days,
                    expired_flag
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (

                    employee_id,
                    selected_employee,
                    str(target_date.date()),
                    grant_days,
                    str(expire_date.date()),
                    0,
                    "未消滅"

                ))

                cursor.execute("""
                INSERT INTO auto_grants (
                    employee_id,
                    grant_year,
                    grant_date
                )
                VALUES (?, ?, ?)
                """, (

                    employee_id,
                    years,
                    str(target_date.date())

                ))

                conn.commit()

                st.success(
                    f"{grant_days}日 自動付与完了"
                )

# =====================================
# 有給消滅処理
# =====================================

st.header("有給消滅チェック")

grant_df = pd.read_sql(
    "SELECT * FROM grant_history",
    conn
)

employee_grants = grant_df[
    grant_df["employee_id"] == employee_id
]

expired_count = 0

today_date = datetime.today().date()

for index, row in employee_grants.iterrows():

    if row["expired_flag"] == "消滅済":

        continue

    expire_date = datetime.strptime(
        row["expire_date"],
        "%Y-%m-%d"
    ).date()

    if today_date >= expire_date:

        remain_expire = (
            float(row["grant_days"])
            - float(row["used_days"])
        )

        if remain_expire < 0:

            remain_expire = 0

        current_leave = employee_leave.iloc[0]

        new_remain = (
            float(current_leave["remain_days"])
            - remain_expire
        )

        if new_remain < 0:

            new_remain = 0

        cursor.execute("""
        UPDATE leave_data
        SET remain_days=?
        WHERE employee_id=?
        """, (

            new_remain,
            employee_id

        ))

        cursor.execute("""
        UPDATE grant_history
        SET expired_flag='消滅済'
        WHERE id=?
        """, (

            int(row["id"]),

        ))

        conn.commit()

        expired_count += 1

if expired_count > 0:

    st.success(
        f"{expired_count}件 消滅処理完了"
    )

else:

    st.info("消滅対象なし")

# =====================================
# 年5日取得義務
# =====================================

st.header("年5日取得義務チェック")

current_year = datetime.today().year

if len(history_df) == 0:

    st.info("有給取得履歴なし")

else:

    history_df["leave_date"] = pd.to_datetime(
        history_df["leave_date"]
    )

    year_history = history_df[
        history_df["leave_date"].dt.year
        == current_year
    ]

    year_summary = year_history.groupby(
        "employee_id"
    )["leave_days"].sum().reset_index()

    warning_list = []

    for index, row in leave_df.iterrows():

        emp_id = row["employee_id"]

        emp_name = row["name"]

        granted = float(row["granted_days"])

        if granted >= 10:

            emp_history = year_summary[
                year_summary["employee_id"]
                == emp_id
            ]

            used_days = 0

            if len(emp_history) > 0:

                used_days = float(
                    emp_history.iloc[0]["leave_days"]
                )

            remain_need = max(
                0,
                5 - used_days
            )

            status = (
                "達成"
                if used_days >= 5
                else "未達"
            )

            warning_list.append({

                "社員ID": emp_id,

                "氏名": emp_name,

                "付与日数": granted,

                "取得日数": used_days,

                "不足日数": remain_need,

                "状態": status

            })

    warning_df = pd.DataFrame(
        warning_list
    )

    if len(warning_df) == 0:

        st.info("対象社員なし")

    else:

        st.dataframe(warning_df)

        not_clear_df = warning_df[
            warning_df["状態"] == "未達"
        ]

        if len(not_clear_df) > 0:

            st.warning(
                f"""
                年5日未達者:
                {len(not_clear_df)}名
                """
            )

        else:

            st.success("全員達成")

# =====================================
# Excel出力
# =====================================

import io
from openpyxl import Workbook

st.header("Excel出力")

# =====================================
# Excel出力
# =====================================

import io

st.header("Excel出力")

if st.button(
    "Excel作成",
    key="excel_output"
):

    # Workbook
    wb = Workbook()

    # シート
    ws = wb.active

    ws.title = "有給台帳"

    # テストデータ
    ws["A1"] = "氏名"
    ws["B1"] = "入社日"
    ws["C1"] = "残有給"

    ws["A2"] = "山田太郎"
    ws["B2"] = "2020-04-01"
    ws["C2"] = 10

    ws["A10"] = "ここが表示されたら成功"

    # メモリ保存
    excel_buffer = io.BytesIO()

    wb.save(excel_buffer)

    excel_buffer.seek(0)

    # ダウンロード
    st.download_button(

        label="Excelダウンロード②",

        data=excel_buffer,

        file_name="有給台帳.xlsx",

        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    



    # =====================================
    # 印刷設定
    # =====================================

    ws.page_setup.paperSize = (
        ws.PAPERSIZE_A4
    )

    ws.page_setup.orientation = (
        ws.ORIENTATION_LANDSCAPE
    )

    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 1

    ws.page_margins = PageMargins(

        left=0.3,
        right=0.3,

        top=0.5,
        bottom=0.5

    )

    ws.sheet_view.zoomScale = 90

    ws.print_options.horizontalCentered = True

    # =====================================
    # タイトル
    # =====================================

    ws["A1"] = "有給休暇管理台帳"

    ws.merge_cells("A1:E1")

    ws["A1"].font = Font(
        bold=True,
        size=18
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    # =====================================
    # 社員情報
    # =====================================

    ws["A3"] = "社員ID"
    ws["B3"] = employee_id

    ws["A4"] = "氏名"
    ws["B4"] = selected_employee

    ws["A5"] = "所属"
    ws["B5"] = employee_row.iloc[0]["branch"]

    ws["A6"] = "入社日"
    ws["B6"] = employee_row.iloc[0]["hire_date"]

    # =====================================
    # 有給情報
    # =====================================

    ws["A8"] = "付与日数"

    ws["B8"] = employee_leave.iloc[0][
        "granted_days"
    ]

    ws["A9"] = "使用日数"

    ws["B9"] = employee_leave.iloc[0][
        "used_days"
    ]

    ws["A10"] = "残日数"

    ws["B10"] = employee_leave.iloc[0][
        "remain_days"
    ]

    # =====================================
    # 履歴
    # =====================================

    ws["A12"] = "取得日"

    ws["B12"] = "取得日数"

    row_no = 13

    for index, row in employee_history.iterrows():

        ws[f"A{row_no}"] = row[
            "leave_date"
        ]

        ws[f"B{row_no}"] = row[
            "leave_days"
        ]

        row_no += 1

    # =====================================
    # 消滅警告
    # =====================================

    ws["D25"] = "消滅警告"

    grant_export_df = pd.read_sql(
        "SELECT * FROM grant_history",
        conn
    )

    employee_grant_export = grant_export_df[

        grant_export_df["employee_id"]
        == employee_id

    ]

    warning_row = 26

    today_date = datetime.today().date()

    for index, row in employee_grant_export.iterrows():

        if row["expired_flag"] == "消滅済":

            continue

        expire_date = datetime.strptime(

            row["expire_date"],

            "%Y-%m-%d"

        ).date()

        remain_days = (

            float(row["grant_days"])
            - float(row["used_days"])

        )

        if remain_days <= 0:

            continue

        days_left = (
            expire_date - today_date
        ).days

        # 30日以内
        if 0 <= days_left <= 30:

            ws[f"D{warning_row}"] = (
                "消滅期限接近"
            )

            ws[f"E{warning_row}"] = (
                str(expire_date)
            )

            ws[f"F{warning_row}"] = (
                remain_days
            )

            # 黄色
            ws[f"D{warning_row}"].fill = yellow_fill
            ws[f"E{warning_row}"].fill = yellow_fill
            ws[f"F{warning_row}"].fill = yellow_fill

            warning_row += 1
   

    # =====================================
    # 枠線
    # =====================================

    thin = Side(
        style="thin",
        color="000000"
    )

    for row in ws.iter_rows(

        min_row=1,

        max_row=row_no,

        min_col=1,

        max_col=5

    ):

        for cell in row:

            cell.border = Border(

                left=thin,
                right=thin,

                top=thin,
                bottom=thin

            )

    # =====================================
    # 列幅
    # =====================================

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # =====================================
    # 月別集計
    # =====================================

    ws["D3"] = "月別出勤集計"

    # attendance 読込
    attendance_export = pd.read_sql(
        "SELECT * FROM attendance",
        conn
    )

    employee_attendance_export = (
        attendance_export[
            attendance_export["employee_id"]
            == employee_id
        ]
    )

    if len(employee_attendance_export) > 0:

        employee_attendance_export[
            "work_date"
        ] = pd.to_datetime(

            employee_attendance_export[
                "work_date"
            ]

        )

        employee_attendance_export["work_date"] = pd.to_datetime(

         employee_attendance_export["work_date"]

)


employee_attendance_export[
            "year_month"
        ] = (

            employee_attendance_export[
                "work_days"
            ].dt.strftime("%Y-%m")

        )

summary_df = employee_attendance_export.groupby(

            [
                "year_month",
                "attendance_type"
            ]

        )["attendance_days"].sum().reset_index()
pivot_df = summary_df.pivot_table(

            index="year_month",

            columns="attendance_type",

            values="attendance_days",

            aggfunc="sum",

            fill_value=0

        ).reset_index()

        # 見出し
ws["D5"] = "年月"
ws["E5"] = "出勤"
ws["F5"] = "有給"
ws["G5"] = "欠勤"

month_row = 6

for index, row in pivot_df.iterrows():

            ws[f"D{month_row}"] = row[
                "year_month"
            ]

            ws[f"E{month_row}"] = row.get(
                "出勤",
                0
            )

            ws[f"F{month_row}"] = row.get(
                "有給",
                0
            )

            ws[f"G{month_row}"] = row.get(
                "欠勤",
                0
            )

            month_row += 1
            

# =====================================
# 8割判定
# =====================================

st.header("8割判定")

if len(employee_attendance) > 0:

    # 出勤扱い日数
    work_days = employee_attendance[
        employee_attendance["attendance_type"].isin([
            "出勤",
            "有給",
            "労災",
            "産休",
            "育休"
        ])
    ]["attendance_days"].sum()

    # 全労働日
    total_days = employee_attendance[
        employee_attendance["attendance_type"].isin([
            "出勤",
            "有給",
            "欠勤"
        ])
    ]["attendance_days"].sum()

    # 出勤率
    attendance_rate = 0

    if total_days > 0:

        attendance_rate = round(
            (work_days / total_days) * 100,
            1
        )

    # 判定表示
    if attendance_rate >= 80:

        st.success(
            f"出勤率: {attendance_rate}% （8割達成）"
        )

    elif attendance_rate >= 70:

        st.warning(
            f"出勤率: {attendance_rate}% （要注意）"
        )

    else:

        st.error(
            f"出勤率: {attendance_rate}% （8割未達）"
        )

else:

    st.info("出勤データなし")

    # =====================================
    # 判定表示
    # =====================================

    # 色設定
red_fill = PatternFill(

        start_color="FF9999",

        end_color="FF9999",

        fill_type="solid"

    )

yellow_fill = PatternFill(

        start_color="FFFF99",

        end_color="FFFF99",

        fill_type="solid"

    )

green_fill = PatternFill(

        start_color="99FF99",

        end_color="99FF99",

        fill_type="solid"

    )

ws["D21"] = "出勤率"

ws["E21"] = f"{attendance_rate}%"

    # 出勤率色変更
if attendance_rate >= 80:

        ws["E21"].fill = green_fill

elif attendance_rate >= 70:

        ws["E21"].fill = yellow_fill

else:

        ws["E21"].fill = red_fill

        ws["D22"] = "判定"

if attendance_rate >= 80:

        ws["E22"] = "有給付与対象"

        ws["E22"].fill = green_fill

else:

        ws["E22"] = "対象外"

        ws["E22"].fill = red_fill


        


    # =====================================
    # 保存
    # =====================================

excel_file = (
        f"{selected_employee}_有給台帳.xlsx"
    )

wb.save(excel_file)

st.success(
        f"{excel_file} 作成完了"
    )

with open(
        excel_file,
        "rb"
    ) as excel_data:

        st.download_button(

    label="Excelダウンロード③",

    data=excel_buffer,

    file_name="有給台帳.xlsx",

    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

)


# =====================================
# PDF出力
# =====================================

st.header("PDF出力")

if st.button("PDF作成"):

    pdf_file = (
        f"{selected_employee}_有給台帳.pdf"
    )

    doc = SimpleDocTemplate(
        pdf_file
    )

    elements = []

    employee_table = Table([

        ["社員ID", employee_id],

        ["氏名", selected_employee],

        ["所属", employee_row.iloc[0]["branch"]],

        ["入社日", employee_row.iloc[0]["hire_date"]]

    ])

    employee_table.setStyle(TableStyle([

        ("GRID", (0, 0), (-1, -1), 1, colors.black)

    ]))

    elements.append(employee_table)

    leave_table = Table([

        ["付与", employee_leave.iloc[0]["granted_days"]],

        ["使用", employee_leave.iloc[0]["used_days"]],

        ["残数", employee_leave.iloc[0]["remain_days"]]

    ])

    leave_table.setStyle(TableStyle([

        ("GRID", (0, 0), (-1, -1), 1, colors.black)

    ]))

    elements.append(leave_table)

    doc.build(elements)

    st.success(
        f"{pdf_file} を作成しました"
    )

    with open(pdf_file, "rb") as pdf_data:

        st.download_button(

            label="PDFダウンロード",

            data=pdf_data,

            file_name=pdf_file,

            mime="application/pdf"

        )

# =====================================
# 全社員Excel出力
# =====================================

if st.session_state["role"] == "admin":

    st.header("全社員Excel出力")

    if st.button(
        "全社員Excel作成",
        key="all_excel_output"
    ):

        all_emp_export = pd.read_sql(
            "SELECT * FROM employees",
            conn
        )

        leave_export = pd.read_sql(
            "SELECT * FROM leave_data",
            conn
        )

        history_export = pd.read_sql(
            "SELECT * FROM leave_history",
            conn
        )

        zip_file_name = (
            "有給台帳一括.zip"
        )

        with zipfile.ZipFile(

            zip_file_name,

            "w",

            zipfile.ZIP_DEFLATED

        ) as zipf:

            for index, emp_row in all_emp_export.iterrows():

                emp_id = emp_row[
                    "employee_id"
                ]

                emp_name = emp_row[
                    "name"
                ]

                emp_leave = leave_export[

                    leave_export[
                        "employee_id"
                    ] == emp_id

                ]

                emp_history = history_export[

                    history_export[
                        "employee_id"
                    ] == emp_id

                ]

                wb = Workbook()

                ws = wb.active

                ws.title = "有給台帳"

                # タイトル
                ws["A1"] = "有給休暇管理台帳"

                ws.merge_cells("A1:E1")

                ws["A1"].font = Font(
                    bold=True,
                    size=18
                )

                # 基本情報
                ws["A3"] = "社員ID"
                ws["B3"] = emp_id

                ws["A4"] = "氏名"
                ws["B4"] = emp_name

                ws["A5"] = "所属"
                ws["B5"] = emp_row[
                    "branch"
                ]

                ws["A6"] = "入社日"
                ws["B6"] = emp_row[
                    "hire_date"
                ]

                # 有給情報
                if len(emp_leave) > 0:

                    ws["A8"] = "付与"
                    ws["B8"] = emp_leave.iloc[0][
                        "granted_days"
                    ]

                    ws["A9"] = "使用"
                    ws["B9"] = emp_leave.iloc[0][
                        "used_days"
                    ]

                    ws["A10"] = "残数"
                    ws["B10"] = emp_leave.iloc[0][
                        "remain_days"
                    ]

                # 履歴
                ws["A12"] = "取得日"
                ws["B12"] = "日数"

                row_no = 13

                for h_index, h_row in emp_history.iterrows():

                    ws[f"A{row_no}"] = h_row[
                        "leave_date"
                    ]

                    ws[f"B{row_no}"] = h_row[
                        "leave_days"
                    ]

                    row_no += 1

                # 列幅
                ws.column_dimensions["A"].width = 20
                ws.column_dimensions["B"].width = 25

                # 保存
                excel_name = (
                    f"{emp_name}_有給台帳.xlsx"
                )

                wb.save(excel_name)

                # =====================================
                # 年別フォルダ
                # =====================================

                current_year = datetime.today().year

                zip_path = (

                    f"{current_year}/"

                    f"{excel_name}"

                )

                zipf.write(

                    excel_name,

                    arcname=zip_path

                )


        st.success(
            "全社員Excel出力完了"
        )

        with open(
            zip_file_name,
            "rb"
        ) as zip_data:

            st.download_button(

                label="ZIPダウンロード",

                data=zip_data,

                file_name=zip_file_name,

                mime="application/zip"

            )


# =====================================
# CSVバックアップ
# =====================================

st.header("CSVバックアップ")

if st.button(
    "バックアップ作成",
    key="backup_create"
):

    # employees
    employees_csv = pd.read_sql(
        "SELECT * FROM employees",
        conn
    )

    employees_csv.to_csv(

        "employees_backup.csv",

        index=False,

        encoding="utf-8-sig"

    )

    # leave_data
    leave_csv = pd.read_sql(
        "SELECT * FROM leave_data",
        conn
    )

    leave_csv.to_csv(

        "leave_data_backup.csv",

        index=False,

        encoding="utf-8-sig"

    )

    # leave_history
    history_csv = pd.read_sql(
        "SELECT * FROM leave_history",
        conn
    )

    history_csv.to_csv(

        "leave_history_backup.csv",

        index=False,

        encoding="utf-8-sig"

    )

    # attendance
    attendance_csv = pd.read_sql(
        "SELECT * FROM attendance",
        conn
    )

    attendance_csv.to_csv(

        "attendance_backup.csv",

        index=False,

        encoding="utf-8-sig"

    )

    # grant_history
    grant_csv = pd.read_sql(
        "SELECT * FROM grant_history",
        conn
    )

    grant_csv.to_csv(

        "grant_history_backup.csv",

        index=False,

        encoding="utf-8-sig"

    )

    st.success(
        "バックアップ完了"
    )

    # =====================================
    # ダウンロード表示
    # =====================================

    backup_files = [

        "employees_backup.csv",

        "leave_data_backup.csv",

        "leave_history_backup.csv",

        "attendance_backup.csv",

        "grant_history_backup.csv"

    ]

    for file_name in backup_files:

        with open(
            file_name,
            "rb"
        ) as file_data:

            st.download_button(

                label=f"{file_name} DL",

                data=file_data,

                file_name=file_name,

                mime="text/csv",

                key=file_name

            )
